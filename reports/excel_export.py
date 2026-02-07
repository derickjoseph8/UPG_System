"""
Excel Export Utility for UPG System Reports

This module provides functions for exporting data to Excel format using openpyxl.
Falls back to CSV if openpyxl is not available.
"""

import csv
from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def create_excel_response(filename, sheets_data, title=None):
    """
    Create an Excel file with multiple sheets.

    Args:
        filename: The filename for download (without extension)
        sheets_data: List of dicts with 'name', 'headers', 'rows' keys
                    Example: [{'name': 'Households', 'headers': ['Name', 'Status'], 'rows': [...]}]
        title: Optional title for the report (shown in first sheet)

    Returns:
        HttpResponse with Excel file, or CSV if openpyxl not available
    """
    if not OPENPYXL_AVAILABLE:
        # Fallback to CSV if openpyxl not installed
        return create_csv_fallback(filename, sheets_data[0] if sheets_data else {'headers': [], 'rows': []})

    wb = Workbook()
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')

    # Remove default sheet if we have data
    if sheets_data:
        default_sheet = wb.active
        wb.remove(default_sheet)

    for idx, sheet_data in enumerate(sheets_data):
        sheet_name = sheet_data.get('name', f'Sheet{idx + 1}')[:31]  # Excel max sheet name length
        ws = wb.create_sheet(title=sheet_name)

        headers = sheet_data.get('headers', [])
        rows = sheet_data.get('rows', [])

        # Add title if provided and this is the first sheet
        start_row = 1
        if title and idx == 0:
            ws.cell(row=1, column=1, value=title)
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))

            # Add export timestamp
            ws.cell(row=2, column=1, value=f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")
            ws.cell(row=2, column=1).font = Font(italic=True, size=10, color='666666')

            start_row = 4

        # Style definitions
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E5A88', end_color='2E5A88', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for row_idx, row in enumerate(rows, start_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

                # Format numbers and dates
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif isinstance(value, datetime):
                    cell.number_format = 'YYYY-MM-DD HH:MM:SS'

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(str(header))

            for row in rows[:100]:  # Check first 100 rows for performance
                if col_idx <= len(row):
                    cell_value = str(row[col_idx - 1]) if row[col_idx - 1] is not None else ''
                    max_length = max(max_length, len(cell_value))

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[col_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}_{timestamp}.xlsx"'

    # Save workbook to response
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    response.write(virtual_workbook.getvalue())

    return response


def create_single_sheet_excel(filename, headers, rows, title=None):
    """
    Convenience function for single-sheet Excel exports.

    Args:
        filename: The filename for download (without extension)
        headers: List of column headers
        rows: List of row data (each row is a list/tuple)
        title: Optional title for the report
    """
    sheets_data = [{'name': 'Report', 'headers': headers, 'rows': rows}]
    return create_excel_response(filename, sheets_data, title)


def create_csv_fallback(filename, sheet_data):
    """
    Fallback to CSV when openpyxl is not available.
    """
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}_{timestamp}.csv"'

    writer = csv.writer(response)
    writer.writerow(sheet_data.get('headers', []))

    for row in sheet_data.get('rows', []):
        writer.writerow(row)

    return response


# ============================================================================
# Pre-built export functions for common report types
# ============================================================================

def export_households_excel(households_queryset, title="Household Report"):
    """Export households to Excel."""
    headers = [
        'Household ID', 'Head of Household', 'National ID', 'Phone Number',
        'Village', 'SubCounty', 'County', 'Status', 'Enrollment Date',
        'Members Count', 'Latitude', 'Longitude'
    ]

    rows = []
    for hh in households_queryset.select_related('village', 'village__subcounty_obj', 'village__subcounty_obj__county'):
        rows.append([
            hh.household_id,
            hh.head_of_household,
            hh.national_id,
            hh.phone_number,
            hh.village.name if hh.village else '',
            hh.village.subcounty_obj.name if hh.village and hh.village.subcounty_obj else '',
            hh.village.subcounty_obj.county.name if hh.village and hh.village.subcounty_obj and hh.village.subcounty_obj.county else '',
            hh.status,
            hh.enrollment_date.strftime('%Y-%m-%d') if hh.enrollment_date else '',
            hh.members.count() if hasattr(hh, 'members') else 0,
            hh.latitude,
            hh.longitude
        ])

    return create_single_sheet_excel('household_report', headers, rows, title)


def export_business_groups_excel(groups_queryset, title="Business Groups Report"):
    """Export business groups to Excel."""
    headers = [
        'Group ID', 'Group Name', 'Registration Number', 'Status',
        'Formation Date', 'Members Count', 'Village', 'SubCounty',
        'Sector', 'Monthly Revenue', 'Total Assets'
    ]

    rows = []
    for group in groups_queryset:
        rows.append([
            group.id,
            group.name,
            group.registration_number or '',
            group.status,
            group.formation_date.strftime('%Y-%m-%d') if group.formation_date else '',
            group.members.count() if hasattr(group, 'members') else 0,
            group.village.name if hasattr(group, 'village') and group.village else '',
            '',  # SubCounty
            group.sector if hasattr(group, 'sector') else '',
            getattr(group, 'monthly_revenue', 0),
            getattr(group, 'total_assets', 0)
        ])

    return create_single_sheet_excel('business_groups_report', headers, rows, title)


def export_savings_groups_excel(groups_queryset, title="Savings Groups Report"):
    """Export savings groups to Excel."""
    headers = [
        'Group ID', 'Group Name', 'Status', 'Formation Date',
        'Members Count', 'Total Savings', 'Total Loans Outstanding',
        'Share Value', 'Meeting Frequency'
    ]

    rows = []
    for group in groups_queryset:
        rows.append([
            group.id,
            group.name,
            group.status if hasattr(group, 'status') else '',
            group.formation_date.strftime('%Y-%m-%d') if hasattr(group, 'formation_date') and group.formation_date else '',
            group.members.count() if hasattr(group, 'members') else 0,
            getattr(group, 'total_savings', 0),
            getattr(group, 'total_loans_outstanding', 0),
            getattr(group, 'share_value', 0),
            getattr(group, 'meeting_frequency', '')
        ])

    return create_single_sheet_excel('savings_groups_report', headers, rows, title)


def export_grants_excel(grants_queryset, title="Grants Report"):
    """Export grants to Excel."""
    headers = [
        'Application ID', 'Grant Type', 'Applicant Name', 'Status',
        'Amount Requested', 'Amount Approved', 'Application Date',
        'Approval Date', 'Village', 'SubCounty'
    ]

    rows = []
    for grant in grants_queryset:
        applicant = getattr(grant, 'applicant', None) or getattr(grant, 'household', None)
        applicant_name = ''
        if applicant:
            if hasattr(applicant, 'head_of_household'):
                applicant_name = applicant.head_of_household
            elif hasattr(applicant, 'name'):
                applicant_name = applicant.name

        rows.append([
            grant.id,
            getattr(grant, 'grant_type', ''),
            applicant_name,
            grant.status if hasattr(grant, 'status') else '',
            getattr(grant, 'amount_requested', 0),
            getattr(grant, 'amount_approved', 0),
            grant.created_at.strftime('%Y-%m-%d') if hasattr(grant, 'created_at') and grant.created_at else '',
            grant.approval_date.strftime('%Y-%m-%d') if hasattr(grant, 'approval_date') and grant.approval_date else '',
            '',  # Village
            ''   # SubCounty
        ])

    return create_single_sheet_excel('grants_report', headers, rows, title)


def export_training_excel(sessions_queryset, title="Training Report"):
    """Export training sessions to Excel."""
    headers = [
        'Session ID', 'Training Module', 'Date', 'Venue',
        'Facilitator', 'Attendees Count', 'Status', 'Duration (hrs)',
        'Village', 'SubCounty'
    ]

    rows = []
    for session in sessions_queryset:
        rows.append([
            session.id,
            session.training_module.name if hasattr(session, 'training_module') and session.training_module else '',
            session.date.strftime('%Y-%m-%d') if hasattr(session, 'date') and session.date else '',
            getattr(session, 'venue', ''),
            session.facilitator.get_full_name() if hasattr(session, 'facilitator') and session.facilitator else '',
            session.attendees.count() if hasattr(session, 'attendees') else 0,
            getattr(session, 'status', ''),
            getattr(session, 'duration_hours', ''),
            getattr(session, 'village', ''),
            ''  # SubCounty
        ])

    return create_single_sheet_excel('training_report', headers, rows, title)


def export_generic_excel(queryset, field_config, filename, title=None):
    """
    Generic export function for any model.

    Args:
        queryset: Django QuerySet
        field_config: List of dicts with 'field' and 'label' keys
                     Example: [{'field': 'name', 'label': 'Name'}, {'field': 'created_at', 'label': 'Created'}]
        filename: Output filename (without extension)
        title: Optional report title
    """
    headers = [fc['label'] for fc in field_config]

    rows = []
    for obj in queryset:
        row = []
        for fc in field_config:
            field_name = fc['field']

            # Handle nested fields (e.g., 'village.name')
            if '.' in field_name:
                parts = field_name.split('.')
                value = obj
                for part in parts:
                    if value is not None:
                        value = getattr(value, part, None)
            else:
                value = getattr(obj, field_name, '')

            # Format value
            if value is None:
                value = ''
            elif isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M:%S')
            elif hasattr(value, 'strftime'):  # date objects
                value = value.strftime('%Y-%m-%d')
            elif callable(value):
                value = value()

            row.append(value)
        rows.append(row)

    return create_single_sheet_excel(filename, headers, rows, title)


# ============================================================================
# Summary Report with Multiple Sheets
# ============================================================================

def export_comprehensive_report(data_dict, title="UPG Comprehensive Report"):
    """
    Create a comprehensive Excel report with multiple sheets.

    Args:
        data_dict: Dict with sheet names as keys, each containing 'headers' and 'rows'
                  Example: {
                      'Households': {'headers': [...], 'rows': [...]},
                      'Business Groups': {'headers': [...], 'rows': [...]}
                  }
        title: Report title
    """
    sheets_data = []

    # Add summary sheet
    summary_headers = ['Metric', 'Count']
    summary_rows = []

    for sheet_name, sheet_info in data_dict.items():
        row_count = len(sheet_info.get('rows', []))
        summary_rows.append([sheet_name, row_count])
        sheets_data.append({
            'name': sheet_name,
            'headers': sheet_info.get('headers', []),
            'rows': sheet_info.get('rows', [])
        })

    # Insert summary as first sheet
    sheets_data.insert(0, {
        'name': 'Summary',
        'headers': summary_headers,
        'rows': summary_rows
    })

    return create_excel_response('comprehensive_report', sheets_data, title)
