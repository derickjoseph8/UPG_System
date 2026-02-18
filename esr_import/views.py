"""
ESR Data Import Views
Comprehensive import for Economic Strengthening Record data
"""
import re
import uuid
import secrets
import string
from datetime import datetime
from io import BytesIO

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db import transaction
from django.db.models import Q
from django.core.paginator import Paginator

from core.decorators import role_required
from core.models import County, AuditLog
from accounts.models import User

from .models import (
    BMCycle, Constituency, District, Division, ESRWard, Location, SubLocation, ESRVillage,
    MentorSupervisor, MentorProfile, MentorAssignment, ESRHousehold, ESRHouseholdMember,
    ESRImportLog
)


def generate_temp_password(length=12):
    """Generate a secure temporary password"""
    alphabet = string.ascii_letters + string.digits + "!@#$%"
    password = ''.join(secrets.choice(alphabet) for _ in range(length))
    # Ensure at least one of each required type
    password = (
        secrets.choice(string.ascii_uppercase) +
        secrets.choice(string.ascii_lowercase) +
        secrets.choice(string.digits) +
        secrets.choice("!@#$%") +
        password[4:]
    )
    return password


def generate_household_code():
    """Generate unique household code"""
    year = datetime.utcnow().year
    unique_part = str(uuid.uuid4().int)[:6].zfill(6)
    return f"HH-{year}-{unique_part}"


def parse_bm_cycle_from_mentor(mentor_value):
    """
    Parse BM cycle prefix from mentor name
    Returns (bm_cycle_name, mentor_name)
    Example: "FY2025/26C1 Jane Doe" -> ("FY2025/26C1", "Jane Doe")
    """
    if not mentor_value:
        return None, None

    mentor_value = str(mentor_value).strip()

    # Pattern to match fiscal year cycle format: FY20XX/XXC# or FY20XX-XXC#
    pattern = r'^(FY\d{4}[/-]\d{2}C\d+)\s+(.+)$'
    match = re.match(pattern, mentor_value, re.IGNORECASE)

    if match:
        bm_cycle = match.group(1).upper()
        mentor_name = match.group(2).strip()
        return bm_cycle, mentor_name

    # If no pattern match, return just the mentor name
    return None, mentor_value


def parse_bm_cycle_details(cycle_name):
    """
    Parse fiscal year and cycle number from BM cycle name
    Example: "FY2025/26C1" -> ("2025/26", 1)
    """
    if not cycle_name:
        return None, None

    pattern = r'^FY(\d{4}[/-]\d{2})C(\d+)$'
    match = re.match(pattern, cycle_name, re.IGNORECASE)

    if match:
        fiscal_year = match.group(1).replace('-', '/')
        cycle_number = int(match.group(2))
        return fiscal_year, cycle_number

    return None, None


def normalize_enum_value(value, choices):
    """Normalize and validate enum value against choices"""
    if not value:
        return ''

    value = str(value).lower().strip().replace(' ', '_').replace('-', '_')

    # Try direct match
    valid_values = [choice[0] for choice in choices]
    if value in valid_values:
        return value

    # Try partial match
    for valid_value in valid_values:
        if value in valid_value or valid_value in value:
            return valid_value

    return ''


@login_required
@role_required(['ict_admin', 'program_manager', 'me_staff'])
def esr_import_page(request):
    """ESR Import main page"""
    import_logs = ESRImportLog.objects.all()[:10]
    context = {
        'import_logs': import_logs,
        'page_title': 'ESR Data Import',
    }
    return render(request, 'esr_import/import.html', context)


@login_required
@role_required(['ict_admin', 'program_manager', 'me_staff'])
def download_template(request):
    """Download ESR import Excel template"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        messages.error(request, "openpyxl is required for Excel operations")
        return redirect('esr_import:import_page')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ESR Data"

    # Define columns
    columns = [
        ("ID_Number", "Unique identifier (National ID) for deduplication", "12345678"),
        ("Mentor_Supervisor", "Name of the mentor supervisor", "John Smith"),
        ("Mentor", "Mentor name (may include BM cycle prefix)", "FY2025/26C1 Jane Doe"),
        ("CountyName", "County name", "Makueni"),
        ("ConstituencyName", "Constituency name", "Makueni Central"),
        ("DistrictName", "District name", "Makueni"),
        ("DivisionName", "Division name", "Wote"),
        ("Ward", "Ward name", "Wote Ward"),
        ("LocationName", "Location name", "Wote Location"),
        ("SubLocationName", "Sub-location name", "Wote Sub-Location"),
        ("Village", "Village name", "Wote Village"),
        ("Landmark", "Landmark near the household", "Near primary school"),
        ("CommunityLeaderName", "Name of the community leader", "Chief Mwangi"),
        ("NumberofhouseholdMembers", "Total number of household members", "5"),
        ("MainCaregiver", "Name of the main caregiver", "Mary Wanjiku"),
        ("NoOfHabitableRooms", "Number of habitable rooms", "3"),
        ("DwellingTenure", "Ownership status (owned/rented/family/squatter/other)", "owned"),
        ("Roof", "Roof material (iron_sheets/tiles/concrete/thatch/mud/other)", "iron_sheets"),
        ("Wall", "Wall material (brick/stone/mud/wood/iron_sheets/other)", "brick"),
        ("Floor", "Floor material (cement/tiles/mud/wood/other)", "cement"),
        ("DwelingUnitRisk", "Risk level (low/medium/high/critical)", "low"),
        ("LightingFuel", "Main lighting source (electricity/solar/kerosene/candle/firewood/none/other)", "solar"),
        ("WaterSource", "Main water source (piped/borehole/well/spring/river/rain/vendor/other)", "borehole"),
        ("HumanwasteDisposal", "Waste disposal method (flush_toilet/pit_latrine/vip_latrine/bush/shared/none/other)", "pit_latrine"),
        ("CookingFuel", "Main cooking fuel (electricity/gas/kerosene/charcoal/firewood/other)", "firewood"),
        ("Latitude", "GPS Latitude (optional)", "-1.8234"),
        ("Longitude", "GPS Longitude (optional)", "37.6283"),
        ("Notes", "Additional notes (optional)", "Household needs additional support"),
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_num, (col_name, description, example) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max(15, len(col_name) + 2)

    # Write example row
    example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    for col_num, (col_name, description, example) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_num, value=example)
        cell.fill = example_fill
        cell.border = thin_border

    # Create instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions.column_dimensions["A"].width = 80

    instructions = [
        "ESR Data Import Template Instructions",
        "",
        "REQUIRED FIELDS:",
        "- ID_Number: Unique national ID for deduplication (REQUIRED)",
        "- Mentor_Supervisor: Name of the supervisor",
        "- Mentor: Can include BM cycle prefix like 'FY2025/26C1 Jane Doe'",
        "- CountyName, ConstituencyName, DistrictName, DivisionName, Ward, LocationName, SubLocationName, Village",
        "- Landmark, CommunityLeaderName, NumberofhouseholdMembers, MainCaregiver, NoOfHabitableRooms",
        "- DwellingTenure, Roof, Wall, Floor, DwelingUnitRisk, LightingFuel, WaterSource, HumanwasteDisposal, CookingFuel",
        "",
        "VALID VALUES FOR CHOICE FIELDS:",
        "DwellingTenure: owned, rented, family, squatter, other",
        "Roof: iron_sheets, tiles, concrete, thatch, mud, other",
        "Wall: brick, stone, mud, wood, iron_sheets, other",
        "Floor: cement, tiles, mud, wood, other",
        "DwelingUnitRisk: low, medium, high, critical",
        "LightingFuel: electricity, solar, kerosene, candle, firewood, none, other",
        "WaterSource: piped, borehole, well, spring, river, rain, vendor, other",
        "HumanwasteDisposal: flush_toilet, pit_latrine, vip_latrine, bush, shared, none, other",
        "CookingFuel: electricity, gas, kerosene, charcoal, firewood, other",
        "",
        "NOTES:",
        "- The BM cycle prefix in the Mentor column will be automatically parsed",
        "- Mentors and supervisors will be created as users with temporary passwords",
        "- Location hierarchy will be auto-created if not exists",
        "- Duplicate records (same ID_Number) will be skipped",
    ]

    for row_idx, instruction in enumerate(instructions, 1):
        ws_instructions.cell(row=row_idx, column=1, value=instruction)

    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ESR_Import_Template_{datetime.now().strftime("%Y%m%d")}.xlsx'

    wb.save(response)
    return response


@login_required
@role_required(['ict_admin', 'program_manager', 'me_staff'])
def process_import(request):
    """Process ESR import from Excel file"""
    if request.method != 'POST':
        return redirect('esr_import:import_page')

    if 'file' not in request.FILES:
        messages.error(request, "Please select a file to import")
        return redirect('esr_import:import_page')

    uploaded_file = request.FILES['file']
    skip_duplicates = request.POST.get('skip_duplicates', 'true') == 'true'

    # Validate file type
    if not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
        messages.error(request, "File must be Excel format (.xlsx or .xls)")
        return redirect('esr_import:import_page')

    try:
        import openpyxl
        workbook = openpyxl.load_workbook(BytesIO(uploaded_file.read()))
        sheet = workbook.active
    except ImportError:
        messages.error(request, "openpyxl library is required for Excel import")
        return redirect('esr_import:import_page')
    except Exception as e:
        messages.error(request, f"Failed to parse Excel file: {str(e)}")
        return redirect('esr_import:import_page')

    # Get headers from first row
    headers = []
    for cell in sheet[1]:
        if cell.value:
            header = str(cell.value).strip().replace(' ', '').replace('_', '').lower()
            headers.append(header)
        else:
            headers.append(f'col_{len(headers)}')

    # Create header mapping
    header_map = {header: idx for idx, header in enumerate(headers)}

    # Initialize tracking
    batch_id = f"ESR-{datetime.now().strftime('%Y%m%d%H%M%S')}-{str(uuid.uuid4())[:8]}"
    stats = {
        'total_records': 0,
        'successful': 0,
        'failed': 0,
        'duplicates': 0,
        'villages_created': 0,
        'mentors_created': 0,
        'supervisors_created': 0,
        'households_created': 0,
        'bm_cycles_created': 0,
    }
    errors = []
    mentor_credentials = []
    supervisor_credentials = []

    # Caches
    county_cache = {}
    constituency_cache = {}
    district_cache = {}
    division_cache = {}
    ward_cache = {}
    location_cache = {}
    sublocation_cache = {}
    village_cache = {}
    bm_cycle_cache = {}
    mentor_cache = {}
    supervisor_cache = {}

    def get_value(row, col_name):
        """Get value from row by column name"""
        normalized = col_name.replace('_', '').replace(' ', '').lower()
        variations = [normalized, col_name.lower(), col_name.lower().replace('_', '')]
        for var in variations:
            if var in header_map:
                idx = header_map[var]
                if idx < len(row):
                    val = row[idx]
                    if val is not None:
                        return str(val).strip() if not isinstance(val, (int, float)) else val
        return None

    with transaction.atomic():
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if all(cell is None for cell in row):
                continue

            stats['total_records'] += 1

            try:
                # Extract ID number
                id_number = get_value(row, 'ID_Number') or get_value(row, 'idnumber')
                if id_number:
                    id_number = str(id_number).strip()
                else:
                    errors.append({"row": row_idx, "error": "Missing required ID_Number"})
                    stats['failed'] += 1
                    continue

                # Check for duplicate
                if skip_duplicates and ESRHousehold.objects.filter(id_number=id_number).exists():
                    stats['duplicates'] += 1
                    continue

                # Extract location hierarchy
                county_name = get_value(row, 'CountyName') or get_value(row, 'county')
                constituency_name = get_value(row, 'ConstituencyName') or get_value(row, 'constituency')
                district_name = get_value(row, 'DistrictName') or get_value(row, 'district')
                division_name = get_value(row, 'DivisionName') or get_value(row, 'division')
                ward_name = get_value(row, 'Ward') or get_value(row, 'wardname')
                location_name = get_value(row, 'LocationName') or get_value(row, 'location')
                sublocation_name = get_value(row, 'SubLocationName') or get_value(row, 'sublocation')
                village_name = get_value(row, 'Village') or get_value(row, 'villagename')

                # Validate required location fields
                if not all([county_name, ward_name, village_name]):
                    errors.append({"row": row_idx, "error": "Missing required location fields"})
                    stats['failed'] += 1
                    continue

                # Create/get County
                county_key = county_name.lower()
                if county_key not in county_cache:
                    county, _ = County.objects.get_or_create(
                        name__iexact=county_name,
                        defaults={'name': county_name}
                    )
                    county_cache[county_key] = county
                county = county_cache[county_key]

                # Create/get Constituency
                constituency = None
                if constituency_name:
                    const_key = f"{county_key}:{constituency_name.lower()}"
                    if const_key not in constituency_cache:
                        constituency, _ = Constituency.objects.get_or_create(
                            name__iexact=constituency_name,
                            county=county,
                            defaults={'name': constituency_name}
                        )
                        constituency_cache[const_key] = constituency
                    constituency = constituency_cache[const_key]

                # Create/get District
                district = None
                if constituency and district_name:
                    dist_key = f"{constituency_name.lower()}:{district_name.lower()}"
                    if dist_key not in district_cache:
                        district, _ = District.objects.get_or_create(
                            name__iexact=district_name,
                            constituency=constituency,
                            defaults={'name': district_name}
                        )
                        district_cache[dist_key] = district
                    district = district_cache[dist_key]

                # Create/get Division
                division = None
                if district and division_name:
                    div_key = f"{district_name.lower()}:{division_name.lower()}"
                    if div_key not in division_cache:
                        division, _ = Division.objects.get_or_create(
                            name__iexact=division_name,
                            district=district,
                            defaults={'name': division_name}
                        )
                        division_cache[div_key] = division
                    division = division_cache[div_key]

                # Create/get Ward
                ward = None
                if division and ward_name:
                    ward_key = f"{division_name.lower()}:{ward_name.lower()}"
                    if ward_key not in ward_cache:
                        ward, _ = ESRWard.objects.get_or_create(
                            name__iexact=ward_name,
                            division=division,
                            defaults={'name': ward_name}
                        )
                        ward_cache[ward_key] = ward
                    ward = ward_cache[ward_key]

                # Create/get Location
                location = None
                if ward and location_name:
                    loc_key = f"{ward_name.lower()}:{location_name.lower()}"
                    if loc_key not in location_cache:
                        location, _ = Location.objects.get_or_create(
                            name__iexact=location_name,
                            ward=ward,
                            defaults={'name': location_name}
                        )
                        location_cache[loc_key] = location
                    location = location_cache[loc_key]

                # Create/get SubLocation
                sublocation = None
                if location and sublocation_name:
                    subloc_key = f"{location_name.lower()}:{sublocation_name.lower()}"
                    if subloc_key not in sublocation_cache:
                        sublocation, _ = SubLocation.objects.get_or_create(
                            name__iexact=sublocation_name,
                            location=location,
                            defaults={'name': sublocation_name}
                        )
                        sublocation_cache[subloc_key] = sublocation
                    sublocation = sublocation_cache[subloc_key]

                # Create/get Village
                village = None
                if sublocation and village_name:
                    village_key = f"{sublocation_name.lower()}:{village_name.lower()}"
                    if village_key not in village_cache:
                        landmark = get_value(row, 'Landmark')
                        leader_name = get_value(row, 'CommunityLeaderName')
                        village, created = ESRVillage.objects.get_or_create(
                            name__iexact=village_name,
                            sub_location=sublocation,
                            defaults={
                                'name': village_name,
                                'landmark': landmark or '',
                                'community_leader_name': leader_name or ''
                            }
                        )
                        if created:
                            stats['villages_created'] += 1
                        village_cache[village_key] = village
                    village = village_cache[village_key]

                if not village:
                    errors.append({"row": row_idx, "error": "Could not create/find village"})
                    stats['failed'] += 1
                    continue

                # Parse mentor and BM cycle
                mentor_value = get_value(row, 'Mentor')
                bm_cycle_name, mentor_name = parse_bm_cycle_from_mentor(mentor_value)

                # Create/get BM Cycle
                bm_cycle = None
                if bm_cycle_name:
                    bm_cycle_key = bm_cycle_name.upper()
                    if bm_cycle_key not in bm_cycle_cache:
                        fiscal_year, cycle_num = parse_bm_cycle_details(bm_cycle_name)
                        bm_cycle, created = BMCycle.objects.get_or_create(
                            name__iexact=bm_cycle_key,
                            defaults={
                                'name': bm_cycle_key,
                                'fiscal_year': fiscal_year or bm_cycle_key,
                                'cycle_number': cycle_num or 1
                            }
                        )
                        if created:
                            stats['bm_cycles_created'] += 1
                        bm_cycle_cache[bm_cycle_key] = bm_cycle
                    bm_cycle = bm_cycle_cache[bm_cycle_key]

                # Create/get Supervisor
                supervisor_name = get_value(row, 'Mentor_Supervisor') or get_value(row, 'mentorsupervisor')
                mentor_supervisor = None
                if supervisor_name:
                    supervisor_key = supervisor_name.lower()
                    if supervisor_key not in supervisor_cache:
                        # Try to find existing
                        try:
                            mentor_supervisor = MentorSupervisor.objects.select_related('user').get(
                                Q(user__first_name__iexact=supervisor_name.split()[0]) |
                                Q(user__username__iexact=supervisor_name.replace(' ', '_').lower())
                            )
                        except MentorSupervisor.DoesNotExist:
                            # Create new supervisor user
                            name_parts = supervisor_name.split(' ', 1)
                            first_name = name_parts[0]
                            last_name = name_parts[1] if len(name_parts) > 1 else name_parts[0]
                            username = f"supervisor_{supervisor_key.replace(' ', '_').replace('.', '_')}"
                            email = f"{username}@esr.local"

                            if not User.objects.filter(Q(email=email) | Q(username=username)).exists():
                                temp_password = generate_temp_password()
                                sup_user = User.objects.create_user(
                                    username=username,
                                    email=email,
                                    password=temp_password,
                                    first_name=first_name,
                                    last_name=last_name,
                                    role='field_associate',
                                    is_active=True
                                )
                                # Mark password as temporary
                                sup_user.set_password(temp_password)
                                sup_user.save()

                                mentor_supervisor = MentorSupervisor.objects.create(
                                    user=sup_user,
                                    is_active=True
                                )
                                stats['supervisors_created'] += 1
                                supervisor_credentials.append({
                                    'name': supervisor_name,
                                    'username': username,
                                    'email': email,
                                    'temp_password': temp_password
                                })

                        if mentor_supervisor:
                            supervisor_cache[supervisor_key] = mentor_supervisor
                    mentor_supervisor = supervisor_cache.get(supervisor_key)

                # Create/get Mentor
                mentor_profile = None
                if mentor_name:
                    mentor_key = mentor_name.lower()
                    if mentor_key not in mentor_cache:
                        # Try to find existing
                        try:
                            mentor_profile = MentorProfile.objects.select_related('user').get(
                                Q(user__first_name__iexact=mentor_name.split()[0]) |
                                Q(user__username__iexact=mentor_name.replace(' ', '_').lower())
                            )
                        except MentorProfile.DoesNotExist:
                            # Create new mentor user
                            name_parts = mentor_name.split(' ', 1)
                            first_name = name_parts[0]
                            last_name = name_parts[1] if len(name_parts) > 1 else name_parts[0]
                            username = f"mentor_{mentor_key.replace(' ', '_').replace('.', '_')}"
                            email = f"{username}@esr.local"

                            if not User.objects.filter(Q(email=email) | Q(username=username)).exists():
                                temp_password = generate_temp_password()
                                mentor_user = User.objects.create_user(
                                    username=username,
                                    email=email,
                                    password=temp_password,
                                    first_name=first_name,
                                    last_name=last_name,
                                    role='mentor',
                                    is_active=True
                                )
                                mentor_user.set_password(temp_password)
                                mentor_user.save()

                                mentor_profile = MentorProfile.objects.create(
                                    user=mentor_user,
                                    supervisor=mentor_supervisor,
                                    profile_completed=False,
                                    is_active=True
                                )
                                stats['mentors_created'] += 1
                                mentor_credentials.append({
                                    'name': mentor_name,
                                    'username': username,
                                    'email': email,
                                    'temp_password': temp_password
                                })

                        if mentor_profile:
                            mentor_cache[mentor_key] = mentor_profile

                            # Create mentor assignment
                            if bm_cycle:
                                MentorAssignment.objects.get_or_create(
                                    mentor=mentor_profile,
                                    village=village,
                                    bm_cycle=bm_cycle,
                                    defaults={'is_active': True}
                                )

                    mentor_profile = mentor_cache.get(mentor_key)

                # Parse household data
                main_caregiver = get_value(row, 'MainCaregiver')
                num_members_val = get_value(row, 'NumberofhouseholdMembers')
                num_members = 1
                if num_members_val:
                    try:
                        num_members = int(float(str(num_members_val)))
                    except (ValueError, TypeError):
                        num_members = 1

                habitable_rooms_val = get_value(row, 'NoOfHabitableRooms')
                habitable_rooms = None
                if habitable_rooms_val:
                    try:
                        habitable_rooms = int(float(str(habitable_rooms_val)))
                    except (ValueError, TypeError):
                        pass

                # Parse enum values
                dwelling_tenure = normalize_enum_value(
                    get_value(row, 'DwellingTenure'),
                    ESRHousehold.DWELLING_TENURE_CHOICES
                )
                roof_type = normalize_enum_value(
                    get_value(row, 'Roof'),
                    ESRHousehold.ROOF_TYPE_CHOICES
                )
                wall_type = normalize_enum_value(
                    get_value(row, 'Wall'),
                    ESRHousehold.WALL_TYPE_CHOICES
                )
                floor_type = normalize_enum_value(
                    get_value(row, 'Floor'),
                    ESRHousehold.FLOOR_TYPE_CHOICES
                )
                dwelling_risk = normalize_enum_value(
                    get_value(row, 'DwelingUnitRisk') or get_value(row, 'dwellingrisk'),
                    ESRHousehold.DWELLING_RISK_CHOICES
                )
                lighting_fuel = normalize_enum_value(
                    get_value(row, 'LightingFuel'),
                    ESRHousehold.LIGHTING_FUEL_CHOICES
                )
                water_source = normalize_enum_value(
                    get_value(row, 'WaterSource'),
                    ESRHousehold.WATER_SOURCE_CHOICES
                )
                waste_disposal = normalize_enum_value(
                    get_value(row, 'HumanwasteDisposal'),
                    ESRHousehold.WASTE_DISPOSAL_CHOICES
                )
                cooking_fuel = normalize_enum_value(
                    get_value(row, 'CookingFuel'),
                    ESRHousehold.COOKING_FUEL_CHOICES
                )

                # Create household
                household = ESRHousehold.objects.create(
                    household_code=generate_household_code(),
                    id_number=id_number,
                    village=village,
                    landmark=get_value(row, 'Landmark') or '',
                    bm_cycle=bm_cycle,
                    main_caregiver=main_caregiver or '',
                    number_of_members=num_members,
                    no_of_habitable_rooms=habitable_rooms,
                    dwelling_tenure=dwelling_tenure,
                    roof_type=roof_type,
                    wall_type=wall_type,
                    floor_type=floor_type,
                    dwelling_risk=dwelling_risk,
                    lighting_fuel=lighting_fuel,
                    water_source=water_source,
                    waste_disposal=waste_disposal,
                    cooking_fuel=cooking_fuel,
                    mentor=mentor_profile,
                    mentor_supervisor=mentor_supervisor,
                    notes=get_value(row, 'Notes') or '',
                    import_batch_id=batch_id,
                    imported_at=timezone.now(),
                    imported_by=request.user
                )

                stats['households_created'] += 1
                stats['successful'] += 1

            except Exception as e:
                errors.append({"row": row_idx, "error": str(e)})
                stats['failed'] += 1

        # Create import log
        import_log = ESRImportLog.objects.create(
            batch_id=batch_id,
            file_name=uploaded_file.name,
            total_records=stats['total_records'],
            successful=stats['successful'],
            failed=stats['failed'],
            duplicates=stats['duplicates'],
            villages_created=stats['villages_created'],
            mentors_created=stats['mentors_created'],
            supervisors_created=stats['supervisors_created'],
            households_created=stats['households_created'],
            bm_cycles_created=stats['bm_cycles_created'],
            status='completed' if stats['failed'] == 0 else 'completed_with_errors',
            error_summary={'errors': errors[:100]} if errors else {},
            mentor_credentials=mentor_credentials,
            supervisor_credentials=supervisor_credentials,
            imported_by=request.user,
            completed_at=timezone.now()
        )

        # Log activity
        AuditLog.objects.create(
            user=request.user,
            action='create',
            model_name='ESRImport',
            object_id=str(import_log.id),
            description=f"Imported {stats['successful']} households from {uploaded_file.name}",
            ip_address=request.META.get('REMOTE_ADDR')
        )

    messages.success(
        request,
        f"Import completed! Created: {stats['households_created']} households, "
        f"{stats['villages_created']} villages, {stats['mentors_created']} mentors, "
        f"{stats['supervisors_created']} supervisors, {stats['bm_cycles_created']} BM cycles. "
        f"Duplicates skipped: {stats['duplicates']}. Errors: {stats['failed']}"
    )

    return redirect('esr_import:import_result', batch_id=batch_id)


@login_required
@role_required(['ict_admin', 'program_manager', 'me_staff'])
def import_result(request, batch_id):
    """View import result details"""
    import_log = get_object_or_404(ESRImportLog, batch_id=batch_id)
    context = {
        'import_log': import_log,
        'page_title': f'Import Result - {batch_id}',
    }
    return render(request, 'esr_import/result.html', context)


@login_required
@role_required(['ict_admin', 'program_manager', 'me_staff'])
def import_history(request):
    """View import history"""
    import_logs = ESRImportLog.objects.all()
    paginator = Paginator(import_logs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'import_logs': page_obj,
        'page_obj': page_obj,
        'page_title': 'Import History',
    }
    return render(request, 'esr_import/history.html', context)


@login_required
@role_required(['ict_admin', 'program_manager', 'me_staff'])
def household_list(request):
    """List ESR households"""
    households = ESRHousehold.objects.select_related('village', 'bm_cycle', 'mentor', 'mentor_supervisor').all()

    # Search
    search = request.GET.get('search', '').strip()
    if search:
        households = households.filter(
            Q(household_code__icontains=search) |
            Q(id_number__icontains=search) |
            Q(main_caregiver__icontains=search)
        )

    # Filter by village
    village_id = request.GET.get('village')
    if village_id:
        households = households.filter(village_id=village_id)

    # Filter by BM cycle
    bm_cycle_id = request.GET.get('bm_cycle')
    if bm_cycle_id:
        households = households.filter(bm_cycle_id=bm_cycle_id)

    paginator = Paginator(households, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'households': page_obj,
        'page_obj': page_obj,
        'villages': ESRVillage.objects.all()[:100],
        'bm_cycles': BMCycle.objects.all(),
        'page_title': 'ESR Households',
    }
    return render(request, 'esr_import/household_list.html', context)


@login_required
@role_required(['ict_admin', 'program_manager', 'me_staff'])
def household_detail(request, household_id):
    """View/Edit ESR household details"""
    household = get_object_or_404(
        ESRHousehold.objects.select_related('village', 'bm_cycle', 'mentor', 'mentor_supervisor'),
        id=household_id
    )

    if request.method == 'POST':
        # Update household
        household.main_caregiver = request.POST.get('main_caregiver', household.main_caregiver)
        household.number_of_members = int(request.POST.get('number_of_members', household.number_of_members) or 1)
        household.no_of_habitable_rooms = int(request.POST.get('no_of_habitable_rooms') or 0) or None
        household.dwelling_tenure = request.POST.get('dwelling_tenure', household.dwelling_tenure)
        household.roof_type = request.POST.get('roof_type', household.roof_type)
        household.wall_type = request.POST.get('wall_type', household.wall_type)
        household.floor_type = request.POST.get('floor_type', household.floor_type)
        household.dwelling_risk = request.POST.get('dwelling_risk', household.dwelling_risk)
        household.lighting_fuel = request.POST.get('lighting_fuel', household.lighting_fuel)
        household.water_source = request.POST.get('water_source', household.water_source)
        household.waste_disposal = request.POST.get('waste_disposal', household.waste_disposal)
        household.cooking_fuel = request.POST.get('cooking_fuel', household.cooking_fuel)
        household.notes = request.POST.get('notes', household.notes)
        household.save()

        messages.success(request, "Household updated successfully")
        return redirect('esr_import:household_detail', household_id=household_id)

    members = household.members.all()

    context = {
        'household': household,
        'members': members,
        'page_title': f'Household - {household.household_code}',
    }
    return render(request, 'esr_import/household_detail.html', context)
