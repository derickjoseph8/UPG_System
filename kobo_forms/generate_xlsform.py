"""
Generate XLSForm Excel file for UPG HH Eligibility Tool
This form includes offline validation using pulldata() with UPG MIS exports
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

def create_xlsform():
    """Create the XLSForm Excel file"""
    wb = openpyxl.Workbook()

    # Create sheets
    survey_sheet = wb.active
    survey_sheet.title = "survey"

    choices_sheet = wb.create_sheet("choices")
    settings_sheet = wb.create_sheet("settings")

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # ==================== SURVEY SHEET ====================
    survey_headers = [
        "type", "name", "label", "required", "constraint", "constraint_message",
        "relevant", "calculation", "appearance", "hint"
    ]

    # Add headers
    for col, header in enumerate(survey_headers, 1):
        cell = survey_sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Survey questions data
    survey_data = [
        # Section 1: Identification
        ["begin_group", "identification", "Section 1: Identification", "", "", "", "", "", "field-list", ""],
        ["select_one bm_cycle_list", "bm_cycle", "Select BM Cycle", "yes", "", "", "", "", "", "Select the Business Mentor Cycle for this survey"],
        ["text", "household_id", "Enter Household ID", "yes", "regex(., '^[A-Z0-9-]+$')", "Invalid ID format. Use uppercase letters, numbers, and hyphens only", "", "", "", "Enter the unique household identifier"],
        ["calculate", "hh_exists", "", "", "", "", "", "pulldata('households', 'hh_id', 'hh_id', ${household_id})", "", ""],
        ["calculate", "hh_name_lookup", "", "", "", "", "", "pulldata('households', 'hh_name', 'hh_id', ${household_id})", "", ""],
        ["calculate", "hh_village_lookup", "", "", "", "", "", "pulldata('households', 'village_name', 'hh_id', ${household_id})", "", ""],
        ["calculate", "hh_ppi_lookup", "", "", "", "", "", "pulldata('households', 'ppi_score', 'hh_id', ${household_id})", "", ""],
        ["calculate", "hh_eligible_lookup", "", "", "", "", "", "pulldata('households', 'is_eligible', 'hh_id', ${household_id})", "", ""],
        ["note", "hh_info_display", "**Household Found:**\nName: ${hh_name_lookup}\nVillage: ${hh_village_lookup}\nPPI Score: ${hh_ppi_lookup}", "", "", "", "${hh_exists} != ''", "", "", ""],
        ["note", "hh_not_found_warning", "**WARNING:** Household ID not found in system! Please verify the ID or register as new household.", "", "", "", "${hh_exists} = '' and ${household_id} != ''", "", "", ""],
        ["text", "household_name", "Household Head Name", "yes", "", "", "", "", "", "Full name of the household head"],
        ["end_group", "", "", "", "", "", "", "", "", ""],

        # Section 2: Consent
        ["begin_group", "consent", "Section 2: Consent", "", "", "", "", "", "field-list", ""],
        ["select_one yes_no", "respondent_consent", "Does the respondent give consent to participate in this survey?", "yes", "", "", "", "", "", "Explain the purpose of the survey before asking for consent"],
        ["note", "no_consent_note", "Survey cannot continue without consent. Thank the respondent and end the survey.", "", "", "", "${respondent_consent} = 'no'", "", "", ""],
        ["end_group", "", "", "", "", "", "", "", "", ""],

        # Section 3: Location
        ["begin_group", "location", "Section 3: Area of Operation", "", "", "", "${respondent_consent} = 'yes'", "", "field-list", ""],
        ["select_one county_list", "county", "Select County", "yes", "", "", "", "", "minimal", ""],
        ["select_one subcounty_list", "subcounty", "Select Sub-County", "yes", "", "", "", "", "minimal", ""],
        ["select_one village_list", "village", "Select Village", "yes", "", "", "", "", "minimal", ""],
        ["geopoint", "household_gps", "Capture Household GPS Location", "yes", "", "", "", "", "", "Stand at the main entrance of the household"],
        ["end_group", "", "", "", "", "", "", "", "", ""],

        # Section 4: Participant Information
        ["begin_group", "participant_info", "Section 4: Participant Basic Information", "", "", "", "${respondent_consent} = 'yes'", "", "field-list", ""],
        ["select_one record_type", "record_type", "Is this a new or existing household?", "yes", "", "", "", "", "", ""],
        ["text", "national_id", "National ID Number", "yes", "regex(., '^[0-9]{7,8}$')", "National ID must be 7 or 8 digits", "", "", "", ""],
        ["select_one gender", "gender", "Gender of participant", "yes", "", "", "", "", "", ""],
        ["date", "date_of_birth", "Date of Birth", "yes", ". <= today()", "Date cannot be in the future", "", "", "", ""],
        ["calculate", "age", "", "", "", "", "", "int((today() - ${date_of_birth}) div 365.25)", "", ""],
        ["note", "age_display", "Age: ${age} years", "", "", "", "", "", "", ""],
        ["select_one yes_no", "owns_phone", "Does participant own a phone/SIM card?", "yes", "", "", "", "", "", ""],
        ["text", "phone_number", "Phone Number", "yes", "regex(., '^(07|01)[0-9]{8}$')", "Enter valid Kenya phone number (07... or 01...)", "${owns_phone} = 'yes'", "", "", ""],
        ["select_one education_level", "highest_education", "Highest level of education in household", "yes", "", "", "", "", "", ""],
        ["select_one education_level", "participant_education", "Participant's education level", "yes", "", "", "", "", "", ""],
        ["select_one relationship", "relationship_to_head", "Relationship to household head", "yes", "", "", "", "", "", ""],
        ["text", "other_relationship", "Specify other relationship", "yes", "", "", "${relationship_to_head} = 'other'", "", "", ""],
        ["end_group", "", "", "", "", "", "", "", "", ""],

        # Section 5: PWD and OVC
        ["begin_group", "pwd_ovc", "Section 5: Disability and Vulnerable Children", "", "", "", "${respondent_consent} = 'yes'", "", "field-list", ""],
        ["select_one yes_no", "has_disability", "Does the participant have any form of disability?", "yes", "", "", "", "", "", ""],
        ["text", "disability_type", "Describe the type of disability", "yes", "", "", "${has_disability} = 'yes'", "", "", ""],
        ["select_one yes_no", "family_disability", "Does any other family member have a disability?", "yes", "", "", "", "", "", ""],
        ["select_one yes_no", "has_ovc", "Are there orphans or vulnerable children in the household?", "yes", "", "", "", "", "", "Children who have lost one or both parents or are in difficult circumstances"],
        ["integer", "ovc_count", "How many OVC in the household?", "yes", ". >= 0 and . <= 20", "Enter a valid number", "${has_ovc} = 'yes'", "", "", ""],
        ["calculate", "pwd_score", "", "", "", "", "", "if(${has_disability} = 'yes', 10, 0) + if(${family_disability} = 'yes', 5, 0) + if(${has_ovc} = 'yes', 5, 0)", "", ""],
        ["end_group", "", "", "", "", "", "", "", "", ""],

        # Section 6: Chronic Illness
        ["begin_group", "chronic_illness", "Section 6: Chronic Illness", "", "", "", "${respondent_consent} = 'yes'", "", "field-list", ""],
        ["select_one yes_no", "on_medication", "Is participant on long-term medication (3+ months)?", "yes", "", "", "", "", "", "For conditions like HIV, diabetes, hypertension, etc."],
        ["select_one yes_no", "family_medication", "Is any family member on long-term medication?", "yes", "", "", "", "", "", ""],
        ["calculate", "chronic_score", "", "", "", "", "", "if(${on_medication} = 'yes', 5, 0) + if(${family_medication} = 'yes', 5, 0)", "", ""],
        ["end_group", "", "", "", "", "", "", "", "", ""],

        # Section 7: Household Composition
        ["begin_group", "hh_composition", "Section 7: Household Details", "", "", "", "${respondent_consent} = 'yes'", "", "field-list", ""],
        ["integer", "total_members", "Total number of household members", "yes", ". > 0 and . <= 30", "Enter between 1 and 30", "", "", "", "Including the respondent"],
        ["integer", "children_under_5", "Number of children under 5 years", "yes", ". >= 0 and . <= ${total_members}", "Cannot exceed total members", "", "", "", ""],
        ["integer", "children_5_to_17", "Number of children aged 5-17 years", "yes", ". >= 0 and . <= ${total_members}", "Cannot exceed total members", "", "", "", ""],
        ["integer", "working_adults", "Number of working-age adults (18-64)", "yes", ". >= 0 and . <= ${total_members}", "Cannot exceed total members", "", "", "", ""],
        ["integer", "elderly_members", "Number of elderly members (65+)", "yes", ". >= 0 and . <= ${total_members}", "Cannot exceed total members", "", "", "", ""],
        ["select_one yes_no", "single_parent", "Is this a single-parent household?", "yes", "", "", "", "", "", "Only one adult caring for children"],
        ["end_group", "", "", "", "", "", "", "", "", ""],

        # Section 8: Savings
        ["begin_group", "savings", "Section 8: Savings", "", "", "", "${respondent_consent} = 'yes'", "", "field-list", ""],
        ["select_one savings_location", "savings_location", "Where do you keep your savings?", "yes", "", "", "", "", "", ""],
        ["integer", "savings_amount", "Approximate total savings amount (KES)", "yes", ". >= 0", "Cannot be negative", "", "", "", "Total across all savings methods"],
        ["select_one yes_no", "in_savings_group", "Are you a member of any savings group?", "yes", "", "", "", "", "", "Chama, merry-go-round, table banking, etc."],
        ["calculate", "savings_score", "", "", "", "", "", "if(${savings_amount} < 1000, 10, if(${savings_amount} < 5000, 5, 0))", "", ""],
        ["end_group", "", "", "", "", "", "", "", "", ""],

        # Section 9: Loans
        ["begin_group", "loans", "Section 9: Loans and Debt", "", "", "", "${respondent_consent} = 'yes'", "", "field-list", ""],
        ["select_one yes_no", "has_loan", "Do you have any outstanding loans?", "yes", "", "", "", "", "", ""],
        ["integer", "loan_amount", "Total outstanding loan amount (KES)", "yes", ". >= 0", "Cannot be negative", "${has_loan} = 'yes'", "", "", ""],
        ["select_one loan_source", "loan_source", "Main source of loan", "yes", "", "", "${has_loan} = 'yes'", "", "", ""],
        ["end_group", "", "", "", "", "", "", "", "", ""],

        # Section 10: Income
        ["begin_group", "income", "Section 10: Household Income", "", "", "", "${respondent_consent} = 'yes'", "", "field-list", ""],
        ["integer", "monthly_income", "Estimated monthly household income (KES)", "yes", ". >= 0", "Cannot be negative", "", "", "", "Total from all sources"],
        ["select_multiple income_sources", "income_sources", "Sources of income (select all that apply)", "yes", "", "", "", "", "", ""],
        ["calculate", "income_score", "", "", "", "", "", "if(${monthly_income} < 3000, 15, if(${monthly_income} < 6000, 10, if(${monthly_income} < 10000, 5, 0)))", "", ""],
        ["end_group", "", "", "", "", "", "", "", "", ""],

        # Section 11: Food Security
        ["begin_group", "food_security", "Section 11: Food Security", "", "", "", "${respondent_consent} = 'yes'", "", "field-list", ""],
        ["select_one meals_count", "meals_per_day", "How many meals does your household typically eat per day?", "yes", "", "", "", "", "", ""],
        ["select_one yes_no", "skipped_meals", "Did your household skip any meals in the last 7 days due to lack of food?", "yes", "", "", "", "", "", ""],
        ["select_one yes_no", "worried_food", "Were you worried about having enough food in the last month?", "yes", "", "", "", "", "", ""],
        ["select_one yes_no", "ate_less_preferred", "Did you eat less preferred foods due to lack of resources?", "yes", "", "", "", "", "", ""],
        ["calculate", "food_score", "", "", "", "", "", "if(${meals_per_day} = 'one', 15, if(${meals_per_day} = 'two', 10, 5)) + if(${skipped_meals} = 'yes', 5, 0) + if(${worried_food} = 'yes', 5, 0)", "", ""],
        ["end_group", "", "", "", "", "", "", "", "", ""],

        # Section 12: Other Programs
        ["begin_group", "other_programs", "Section 12: Other Programs", "", "", "", "${respondent_consent} = 'yes'", "", "field-list", ""],
        ["select_one yes_no", "in_graduation_program", "Is participant currently in another graduation/livelihood program?", "yes", "", "", "", "", "", "Programs like GiveDirectly, BRAC, etc."],
        ["text", "program_name", "Name of the program", "yes", "", "", "${in_graduation_program} = 'yes'", "", "", ""],
        ["select_one program_status", "program_status", "Current status in that program", "yes", "", "", "${in_graduation_program} = 'yes'", "", "", ""],
        ["calculate", "program_score", "", "", "", "", "", "if(${in_graduation_program} = 'yes' and ${program_status} = 'active', -20, 0)", "", ""],
        ["end_group", "", "", "", "", "", "", "", "", ""],

        # Section 13: Willingness
        ["begin_group", "willingness", "Section 13: Willingness to Participate", "", "", "", "${respondent_consent} = 'yes'", "", "field-list", ""],
        ["select_one yes_no", "willing_to_join", "Is participant willing to join the UPG program?", "yes", "", "", "", "", "", ""],
        ["select_one yes_no", "can_attend_training", "Can participant attend weekly training sessions?", "yes", "", "", "${willing_to_join} = 'yes'", "", "", "Training is typically 2-3 hours per week"],
        ["select_one yes_no", "willing_form_group", "Is participant willing to form a business group with others?", "yes", "", "", "${willing_to_join} = 'yes'", "", "", "Business groups typically have 2-3 members"],
        ["select_one yes_no", "willing_save", "Is participant willing to save regularly?", "yes", "", "", "${willing_to_join} = 'yes'", "", "", ""],
        ["end_group", "", "", "", "", "", "", "", "", ""],

        # Section 14: Scoring
        ["begin_group", "scoring", "Section 14: Eligibility Assessment", "", "", "", "${respondent_consent} = 'yes'", "", "field-list", ""],
        ["calculate", "total_score", "", "", "", "", "", "${pwd_score} + ${chronic_score} + ${savings_score} + ${income_score} + ${food_score} + ${program_score}", "", ""],
        ["note", "score_display", "**TOTAL ELIGIBILITY SCORE: ${total_score}**", "", "", "", "", "", "", ""],
        ["calculate", "eligibility_status", "", "", "", "", "", "if(${total_score} >= 40 and ${willing_to_join} = 'yes' and (${in_graduation_program} = 'no' or ${program_status} != 'active'), 'eligible', 'not_eligible')", "", ""],
        ["note", "eligible_note", "**RESULT: ELIGIBLE FOR UPG PROGRAM**\n\nThis household meets the eligibility criteria.", "", "", "", "${eligibility_status} = 'eligible'", "", "", ""],
        ["note", "not_eligible_note", "**RESULT: NOT ELIGIBLE FOR UPG PROGRAM**\n\nThis household does not meet the eligibility criteria at this time.", "", "", "", "${eligibility_status} = 'not_eligible'", "", "", ""],
        ["end_group", "", "", "", "", "", "", "", "", ""],

        # Section 15: Confirmation
        ["begin_group", "confirmation", "Section 15: Survey Confirmation", "", "", "", "${respondent_consent} = 'yes'", "", "", ""],
        ["image", "signature", "Capture participant signature or thumbprint", "yes", "", "", "", "", "signature", ""],
        ["text", "surveyor_name", "Surveyor Name", "yes", "", "", "", "", "", ""],
        ["text", "surveyor_id", "Surveyor ID/Code", "yes", "", "", "", "", "", ""],
        ["dateTime", "survey_datetime", "Survey Date and Time", "yes", "", "", "", "", "", ""],
        ["text", "comments", "Additional Comments/Observations", "", "", "", "", "", "multiline", ""],
        ["end_group", "", "", "", "", "", "", "", "", ""],
    ]

    # Write survey data
    for row_num, row_data in enumerate(survey_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = survey_sheet.cell(row=row_num, column=col_num, value=value)
            # Highlight section rows
            if row_data[0] in ["begin_group", "end_group"]:
                cell.fill = section_fill

    # Set column widths for survey sheet
    survey_sheet.column_dimensions['A'].width = 25
    survey_sheet.column_dimensions['B'].width = 30
    survey_sheet.column_dimensions['C'].width = 60
    survey_sheet.column_dimensions['D'].width = 10
    survey_sheet.column_dimensions['E'].width = 40
    survey_sheet.column_dimensions['F'].width = 50
    survey_sheet.column_dimensions['G'].width = 50
    survey_sheet.column_dimensions['H'].width = 70
    survey_sheet.column_dimensions['I'].width = 15
    survey_sheet.column_dimensions['J'].width = 50

    # ==================== CHOICES SHEET ====================
    choices_headers = ["list_name", "name", "label"]
    for col, header in enumerate(choices_headers, 1):
        cell = choices_sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    choices_data = [
        # Yes/No
        ["yes_no", "yes", "Yes"],
        ["yes_no", "no", "No"],

        # Gender
        ["gender", "male", "Male"],
        ["gender", "female", "Female"],

        # Record type
        ["record_type", "new", "New Household"],
        ["record_type", "existing", "Existing Household"],

        # Education level
        ["education_level", "none", "No formal education"],
        ["education_level", "primary_incomplete", "Primary (incomplete)"],
        ["education_level", "primary_complete", "Primary (complete)"],
        ["education_level", "secondary_incomplete", "Secondary (incomplete)"],
        ["education_level", "secondary_complete", "Secondary (complete)"],
        ["education_level", "tertiary", "Tertiary/College/University"],

        # Relationship
        ["relationship", "head", "Household Head"],
        ["relationship", "spouse", "Spouse"],
        ["relationship", "child", "Child"],
        ["relationship", "parent", "Parent"],
        ["relationship", "grandparent", "Grandparent"],
        ["relationship", "sibling", "Sibling"],
        ["relationship", "other_relative", "Other Relative"],
        ["relationship", "other", "Other (specify)"],

        # Savings location
        ["savings_location", "home", "At home"],
        ["savings_location", "bank", "Bank account"],
        ["savings_location", "mobile_money", "Mobile money (M-Pesa)"],
        ["savings_location", "sacco", "SACCO"],
        ["savings_location", "savings_group", "Savings group/Chama"],
        ["savings_location", "none", "No savings"],

        # Loan source
        ["loan_source", "bank", "Bank"],
        ["loan_source", "mobile_loan", "Mobile loan (Fuliza, Tala, Branch)"],
        ["loan_source", "sacco", "SACCO"],
        ["loan_source", "savings_group", "Savings group"],
        ["loan_source", "family", "Family/Friends"],
        ["loan_source", "shylock", "Shylock/Money lender"],
        ["loan_source", "other", "Other"],

        # Income sources
        ["income_sources", "farming", "Farming/Agriculture"],
        ["income_sources", "livestock", "Livestock keeping"],
        ["income_sources", "casual_labor", "Casual labor"],
        ["income_sources", "small_business", "Small business/Trade"],
        ["income_sources", "formal_employment", "Formal employment"],
        ["income_sources", "remittances", "Remittances"],
        ["income_sources", "pension", "Pension"],
        ["income_sources", "government_support", "Government support (cash transfer)"],
        ["income_sources", "none", "No regular income"],

        # Meals count
        ["meals_count", "one", "One meal per day"],
        ["meals_count", "two", "Two meals per day"],
        ["meals_count", "three_plus", "Three or more meals per day"],

        # Program status
        ["program_status", "active", "Currently active"],
        ["program_status", "graduated", "Graduated/Completed"],
        ["program_status", "dropped", "Dropped out"],
        ["program_status", "pending", "Pending/Enrolled but not started"],

        # BM Cycle list (placeholder - will be populated from CSV)
        ["bm_cycle_list", "cycle_1", "FY25 Cycle 1"],
        ["bm_cycle_list", "cycle_2", "FY25 Cycle 2"],
        ["bm_cycle_list", "cycle_3", "FY25 Cycle 3"],
        ["bm_cycle_list", "cycle_4", "FY25 Cycle 4"],

        # County list (placeholder - will use pulldata from CSV)
        ["county_list", "kilifi", "Kilifi"],
        ["county_list", "kwale", "Kwale"],
        ["county_list", "mombasa", "Mombasa"],

        # Sub-county list (placeholder)
        ["subcounty_list", "kaloleni", "Kaloleni"],
        ["subcounty_list", "rabai", "Rabai"],
        ["subcounty_list", "ganze", "Ganze"],

        # Village list (placeholder)
        ["village_list", "village_1", "Village 1"],
        ["village_list", "village_2", "Village 2"],
    ]

    for row_num, row_data in enumerate(choices_data, 2):
        for col_num, value in enumerate(row_data, 1):
            choices_sheet.cell(row=row_num, column=col_num, value=value)

    # Set column widths
    choices_sheet.column_dimensions['A'].width = 20
    choices_sheet.column_dimensions['B'].width = 25
    choices_sheet.column_dimensions['C'].width = 40

    # ==================== SETTINGS SHEET ====================
    settings_headers = ["form_title", "form_id", "version", "style", "default_language"]
    for col, header in enumerate(settings_headers, 1):
        cell = settings_sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    settings_data = [
        "UPG Household Eligibility Tool",
        "upg_hh_eligibility_v1",
        datetime.now().strftime("%Y.%m.%d"),
        "pages",
        "English (en)"
    ]

    for col, value in enumerate(settings_data, 1):
        settings_sheet.cell(row=2, column=col, value=value)

    # Set column widths
    settings_sheet.column_dimensions['A'].width = 35
    settings_sheet.column_dimensions['B'].width = 25
    settings_sheet.column_dimensions['C'].width = 15
    settings_sheet.column_dimensions['D'].width = 10
    settings_sheet.column_dimensions['E'].width = 20

    return wb


def main():
    """Generate and save the XLSForm"""
    # Get the directory of this script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "UPG_HH_Eligibility_Tool.xlsx")

    wb = create_xlsform()
    wb.save(output_path)

    print(f"XLSForm created successfully: {output_path}")
    print("\nNext steps:")
    print("1. Open KoBoToolbox (https://kf.kobotoolbox.org)")
    print("2. Create new project > Import XLSForm")
    print("3. Upload this Excel file")
    print("4. Go to Settings > Media Files")
    print("5. Upload CSV files from UPG MIS (households.csv, bm_cycles.csv, villages.csv)")
    print("6. Deploy the form to collect data offline")


if __name__ == "__main__":
    main()
